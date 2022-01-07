import datetime
import pathlib
import time
import zipfile
from typing import Optional
import sys

import pytz
from openpyxl import load_workbook

DATE_FMT = '%Y-%m-%d %H:%M:%S'

epoch = datetime.datetime(1970, 1, 1, tzinfo=datetime.timezone.utc)


def timestamp_ns(time_point: Optional[datetime.datetime] = None) -> Optional[int]:
    if time_point is None:
        time_point = datetime.datetime.now(datetime.timezone.utc)

    return int((time_point - epoch).total_seconds()) * 1000_000_000


def load_zip(path: pathlib.Path) -> None:
    zip_start = time.monotonic_ns()
    zipf = zipfile.ZipFile(path)
    for zipinfo in zipf.infolist():
        with zipf.open(zipinfo.filename, 'r') as excel_buf:
            wb_start = time.monotonic_ns()
            process_workbook(excel_buf)
            wb_end = time.monotonic_ns()
            print(
                f'process_workbook,file_name={zipinfo.filename} size={zipinfo.file_size},time={wb_end - wb_start} {timestamp_ns()}', file=sys.stderr)
            mtime = zipinfo.date_time
    zip_end = time.monotonic_ns()
    file_date = datetime.datetime(*mtime, tzinfo=datetime.timezone.utc)
    print(
        f'process_zipfile file_date="{file_date.strftime(DATE_FMT)}",size={path.stat().st_size},time={zip_end - zip_start} {timestamp_ns()}', file=sys.stderr)


tag_keys = [
    'Id',
    'Country',
    'Loan Originator',
    'Mintos Risk',
    'Loan Type',
    'Term',
    'Collateral',
    'Initial LTV',
    'LTV',
    'Loan Status',
    'Currency',
    'Buyback',
    'Extendable schedule',
    'Loan Originator Status',
    'In Recovery'
]

skip = [
]


def process_workbook(excel_buf):
    workbook = load_workbook(excel_buf, read_only=True, data_only=True)
    try:
        ws = workbook.active

        header = next(ws.iter_rows(1, 1, values_only=True))

        for row in ws.iter_rows(2, max_col=len(header), values_only=True):
            values = list(row)
            for i in [1, 2, 3]:
                values[i] = parse_date(values[i])

            timestamp = timestamp_ns(values[1] or values[2])

            for i in [1, 2, 3]:
                values[i] = values[i].strftime(DATE_FMT+" %Z") if values[i] else None

            for i in [ 10, 18, 19, 21]:
                values[i] = yes_no_bool(values[i])

            for i in range(len(values)):
                if header[i] not in tag_keys:
                    values[i] = quote(values[i])

            tags = [f'{escape(i[0])}={escape(i[1])}' for i in
                    sorted(zip(header, values), key=lambda i: i[0]) if i[0] in tag_keys]

            fields = [f'{escape(i[0])}={i[1]}' for i in
                      (zip(header, values)) if i[0] not in tag_keys and i[0] not in skip and i[1]]

            line = 'loan,' + ','.join(tags) + ' ' + ','.join(fields) + ' ' + str(timestamp)

            print(line)
    finally:
        workbook.close()


def escape(obj: any) -> any:
    if type(obj) is str:
        return obj.replace(' ', '\\ ')
    else:
        return obj


def yes_no_bool(s: Optional[str]) -> Optional[bool]:
    if s is None:
        return None
    elif s.lower() == 'yes':
        return True
    else:
        return False


def parse_date(s: Optional[str]) -> Optional[datetime.datetime]:
    if s is None:
        return None

    return datetime.datetime.strptime(s, DATE_FMT).astimezone(pytz.timezone('Europe/Riga'))


def quote(obj):
    if type(obj) is str:
        return f'"{obj}"'
    else:
        return obj


def influx_int(i: Optional[int]) -> Optional[str]:
    if i is None:
        return None
    else:
        return str(i) + 'i'
